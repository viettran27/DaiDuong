import pandas as pd
from fastapi import HTTPException
from sqlalchemy import inspect
from io import BytesIO
from fastapi.responses import StreamingResponse
from pandas import DataFrame
from sqlalchemy.engine import Engine
from sqlalchemy import text
from app.helper.utils import is_date, is_number, is_integer
from datetime import datetime, time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, NamedStyle

def export_columns(table_name: str, fileName, engine: Engine, examples: list[list], expected_columns: list[str] = None):
    try:
        inspector = inspect(engine)
        columns = inspector.get_columns(table_name)
        
        column_names = [col["name"] for col in columns]
        if expected_columns:
            column_names = [col for col in column_names if col not in expected_columns]
                
        df = pd.DataFrame(data=examples, columns=column_names)

        thin_border = Border(left=Side(style='none'),
                        right=Side(style='none'),
                        top=Side(style='none'),
                        bottom=Side(style='none'))
        font = Font(bold=True)
        date_format = NamedStyle(name="date", number_format="DD/MM/YYYY")

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            dateCol = ["WorkDate", "NGAY", "NGAY_DI_HO_TRO"]
            
            ws = writer.sheets['Sheet1']
            for i, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(i)
                cell = ws[f'{col_letter}1']
                cell.font = font
                cell.border = thin_border
                max_length = max(len(str(value)) for value in [column] + df[column].astype(str).tolist()) + 5
                ws.column_dimensions[col_letter].width = max_length

                if column in dateCol: 
                    workdate_col = i

            for row in ws.iter_rows(min_row=2, min_col=workdate_col, max_col=workdate_col):
                for cell in row:
                    cell.style = date_format

        output.seek(0)

        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fileName}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
def execute_filtered_query(db, table_name, order_by, filters, page=1, size=20):
    offset = (page - 1) * size
    query = f"SELECT * FROM {table_name}"
    
    conditions = []
    for column, value in filters.items():
        if value is not None:
            conditions.append(f"{column} = '{value}'")
    
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    
    query += f" ORDER BY {order_by} DESC OFFSET {offset} ROWS FETCH NEXT {size} ROWS ONLY"
    
    res = db.execute(text(query)).fetchall()
    data = [dict(row._mapping) for row in res]
    
    count_query = f"SELECT COUNT(*) FROM {table_name}"
    if conditions:
        count_query += " WHERE " + " AND ".join(conditions)
    total_count = db.execute(text(count_query)).scalar()
    
    return {
        "data": data,
        "total_page": (total_count + size - 1) // size
    }

def import_to_sql(df: DataFrame, table_name: str, dtype: dict, engine: Engine):
    try:
        with engine.connect() as connection:
            df.to_sql(name=table_name, con=connection, if_exists="append", index=False, dtype=dtype)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
def validate_and_import_excel(file, engine, table_name, dtype, columns_to_validate, dateColumn = None):
    # Kiểm tra định dạng file
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Hãy dùng file Excel")
    
    # Đọc dữ liệu từ Excel
    df = pd.read_excel(file.file)

    invalid_data = {col: [] for col in columns_to_validate}

    # Kiểm tra dữ liệu cho từng cột
    for idx in range(len(df)):
        for col, col_type in columns_to_validate.items():            
            value = df.at[idx, col]
            if pd.isnull(value):
                continue
            elif col_type == 'date' and not is_date(value):
                invalid_data[col].append(idx + 2)
            elif col_type == 'number' and not is_number(value):
                invalid_data[col].append(idx + 2)
            elif col_type == 'integer' and not is_integer(value):
                invalid_data[col].append(idx + 2)
            elif col_type == 'time':
                if dateColumn and pd.notnull(value):
                    if isinstance(value, str):
                        value = datetime.strptime(value, "%H:%M").time()
                    date_value = df.at[idx, dateColumn]
                    if pd.notnull(date_value):
                        df.at[idx, col] = datetime.combine(date_value, value)
                    else:
                        invalid_data[col].append(idx + 2)

    # Ném lỗi nếu có dữ liệu không hợp lệ
    for col, invalid_rows in invalid_data.items():
        if invalid_rows:
            raise HTTPException(status_code=400, detail=f"{col} sai định dạng ở dòng {', '.join(map(str, invalid_rows))}")
        
    # Nhập dữ liệu vào SQL Server
    import_to_sql(df, table_name, dtype, engine)

    return {
        "message": "Nhập dữ liệu thành công"
    }

        